# -*- coding: utf-8 -*-
"""
Generate LTV Cohort Analysis (Assignment 6)

- Cohort definition: customer's first purchase quarter (YYYY-Q#)
- Metric: LTV = cumulative revenue (Sales) per acquired customer in cohort
- Two views:
  1) By age years since cohort start (Y0, Y1, Y2, ...)
  2) By calendar year (cumulative to each year)
- Output: Excel with heatmaps

Usage (optional):
  python generate_ltv_cohort.py --input "../goit_pa_hm_3/Sample - Superstore.xls" --out "Fefelov_PA_assignment_6.xlsx"

The script will try sensible defaults if arguments are not provided.
"""

from __future__ import annotations
import argparse
import os
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns

# Excel formatting
from openpyxl.styles import Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule

# -----------------------------
# Helpers
# -----------------------------

def find_default_input(base_dir: Path) -> Path:
    """Try to find the Superstore dataset from Assignment 3."""
    candidates = [
        base_dir / "goit_pa_hm_3" / "Sample - Superstore.xls",
        base_dir / "goit_pa_hm_3" / "Sample - Superstore.xlsx",
        base_dir / "Sample - Superstore.xls",
        base_dir / "Sample - Superstore.xlsx",
    ]
    for p in candidates:
        if p.exists():
            return p
    raise FileNotFoundError("Could not find 'Sample - Superstore' file. Provide --input path.")


def load_data(path: Path) -> pd.DataFrame:
    # Force xlrd engine for legacy .xls files
    if path.suffix.lower() == ".xls":
        df = pd.read_excel(path, engine="xlrd")
    else:
        df = pd.read_excel(path)
    # Required columns sanity check
    required = {"Order Date", "Customer ID", "Sales"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"Missing required columns: {missing}")
    df["Order Date"] = pd.to_datetime(df["Order Date"])  # normalize
    return df


def prepare_cohorts(df: pd.DataFrame) -> pd.DataFrame:
    # First purchase per customer -> cohort
    first_order = df.groupby("Customer ID")["Order Date"].min().rename("FirstOrderDate")
    df = df.merge(first_order, on="Customer ID", how="left")

    # Cohort quarter label like '2014-Q1'
    cohort_period = df["FirstOrderDate"].dt.to_period("Q")
    df["Cohort_Quarter"] = cohort_period.dt.strftime("%Y-Q%q")

    # Components for age calculations
    df["Order_Year"] = df["Order Date"].dt.year
    df["Cohort_Start_Year"] = df["FirstOrderDate"].dt.year
    df["Age_Year"] = df["Order_Year"] - df["Cohort_Start_Year"]
    df = df[df["Age_Year"] >= 0]  # guard
    return df


def ltv_by_age(df: pd.DataFrame) -> pd.DataFrame:
    # Cohort size (acquired customers)
    cohort_sizes = df.groupby("Cohort_Quarter")["Customer ID"].nunique().rename("Cohort_Size")

    # Revenue by cohort and age year
    rev = df.groupby(["Cohort_Quarter", "Age_Year"])['Sales'].sum().reset_index()

    # Pivot to wide and compute cumulative across age (Y0..Yn)
    pivot = rev.pivot(index="Cohort_Quarter", columns="Age_Year", values="Sales").fillna(0).sort_index()
    pivot = pivot.reindex(columns=sorted(pivot.columns))  # ensure Y0..Yn order

    # Average LTV per acquired customer
    pivot = pivot.div(cohort_sizes, axis=0).fillna(0)

    # Cumulative along age
    pivot_cum = pivot.cumsum(axis=1)
    # Rename columns to 'Y0','Y1',...
    pivot_cum.columns = [f"Y{int(c)}" for c in pivot_cum.columns]

    # Add cohort size as first column for reference
    out = cohort_sizes.to_frame().join(pivot_cum, how="left").fillna(0)
    # Order columns: Cohort_Size, Y0..Yn
    y_cols = [c for c in out.columns if c.startswith("Y")]
    return out[["Cohort_Size", *y_cols]].round(2)


def ltv_by_calendar_year(df: pd.DataFrame) -> pd.DataFrame:
    cohort_sizes = df.groupby("Cohort_Quarter")["Customer ID"].nunique().rename("Cohort_Size")
    rev = df.groupby(["Cohort_Quarter", "Order_Year"])['Sales'].sum().reset_index()

    pivot = rev.pivot(index="Cohort_Quarter", columns="Order_Year", values="Sales").fillna(0).sort_index()
    pivot = pivot.reindex(columns=sorted(pivot.columns))

    pivot = pivot.div(cohort_sizes, axis=0).fillna(0)
    pivot_cum = pivot.cumsum(axis=1)

    out = cohort_sizes.to_frame().join(pivot_cum, how="left").fillna(0)
    # Order columns: Cohort_Size, years ascending
    year_cols = [c for c in out.columns if isinstance(c, (int, np.integer))]
    return out[["Cohort_Size", *sorted(year_cols)]].round(2)


def write_excel(out_path: Path, by_age: pd.DataFrame, by_year: pd.DataFrame) -> None:
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # Overview sheet
        overview = pd.DataFrame(
            {
                "Item": [
                    "Definition",
                    "Metric",
                    "Note",
                ],
                "Value": [
                    "Cohort = first purchase quarter (YYYY-Q#)",
                    "LTV = cumulative Sales per acquired customer",
                    "Two heatmaps: age years and calendar years",
                ],
            }
        )
        overview.to_excel(writer, sheet_name="README", index=False)

        by_age.to_excel(writer, sheet_name="Cohort_LTV_by_Age", index=True)
        by_year.to_excel(writer, sheet_name="Cohort_LTV_by_Year", index=True)

    # Apply heatmaps with openpyxl
    from openpyxl import load_workbook

    wb = load_workbook(out_path)

    def format_heatmap(ws_name: str):
        ws = wb[ws_name]
        # Find bounds
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 2 or max_col < 3:
            return
        # Data range excludes index header column A and the Cohort_Size column B
        start_col = 3  # column C
        start_row = 2  # skip header
        data_range = ws.cell(row=start_row, column=start_col).coordinate + ":" + ws.cell(
            row=max_row, column=max_col
        ).coordinate

        # 3-color scale heatmap
        rule = ColorScaleRule(
            start_type="min", start_color="F5FBFF",
            mid_type="percentile", mid_value=50, mid_color="9AD1F5",
            end_type="max", end_color="004E98",
        )
        ws.conditional_formatting.add(data_range, rule)

        # Currency format and nicer header
        # Set currency for data and Cohort_Size column (B)
        for r in range(2, max_row + 1):
            # Cohort_Size (integer)
            ws.cell(row=r, column=2).number_format = "#,##0"
            for c in range(3, max_col + 1):
                ws.cell(row=r, column=c).number_format = "[$$-409]#,##0.00"

        # Style header row
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")

        # Freeze panes at B2
        ws.freeze_panes = ws["B2"]

        # Adjust widths a bit
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 12
        for c in range(3, max_col + 1):
            ws.column_dimensions[chr(64 + c if c <= 26 else 64)].width = 12

    format_heatmap("Cohort_LTV_by_Age")
    format_heatmap("Cohort_LTV_by_Year")

    wb.save(out_path)


def save_png_heatmaps(out_dir: Path, by_age: pd.DataFrame, by_year: pd.DataFrame) -> None:
    """Save quick PNG heatmaps for visual reference (optional for LMS)."""
    sns.set_theme(style="white")

    # By age (drop cohort size)
    age_tbl = by_age.drop(columns=["Cohort_Size"], errors="ignore").copy()
    # Ensure numeric
    age_tbl = age_tbl.apply(pd.to_numeric, errors="coerce").fillna(0)

    plt.figure(figsize=(12, max(4, len(age_tbl) * 0.3)))
    ax = sns.heatmap(age_tbl, cmap="Blues", annot=False, cbar=True)
    ax.set_title("Cohort LTV by Age (cumulative $ per customer)")
    plt.tight_layout()
    age_png = out_dir / "Fefelov_PA_assignment_6_by_age.png"
    plt.savefig(age_png, dpi=200, bbox_inches="tight")
    plt.close()

    # By calendar year (drop cohort size)
    year_tbl = by_year.drop(columns=["Cohort_Size"], errors="ignore").copy()
    year_tbl = year_tbl.apply(pd.to_numeric, errors="coerce").fillna(0)

    plt.figure(figsize=(12, max(4, len(year_tbl) * 0.3)))
    ax = sns.heatmap(year_tbl, cmap="Blues", annot=False, cbar=True)
    ax.set_title("Cohort LTV by Calendar Year (cumulative $ per customer)")
    plt.tight_layout()
    year_png = out_dir / "Fefelov_PA_assignment_6_by_year.png"
    plt.savefig(year_png, dpi=200, bbox_inches="tight")
    plt.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=str, default=None, help="Path to Superstore dataset (.xls/.xlsx)")
    parser.add_argument("--out", type=str, default="Fefelov_PA_assignment_6.xlsx", help="Output Excel path")
    args = parser.parse_args()

    base_dir = Path(__file__).resolve().parents[1]  # repo root (Product_Analytics)

    input_path = Path(args.input) if args.input else find_default_input(base_dir)
    output_path = Path(args.out)
    if not output_path.is_absolute():
        # default to script directory
        output_path = Path(__file__).resolve().parent / output_path

    print(f"Loading dataset from: {input_path}")
    df = load_data(input_path)

    print("Preparing cohorts...")
    df = prepare_cohorts(df)

    print("Computing LTV by age years...")
    by_age = ltv_by_age(df)

    print("Computing LTV by calendar year...")
    by_year = ltv_by_calendar_year(df)

    print(f"Writing Excel to: {output_path}")
    write_excel(output_path, by_age, by_year)
    # Optional: also export PNG heatmaps for quick preview
    try:
        save_png_heatmaps(output_path.parent, by_age, by_year)
        print("Saved PNG heatmaps next to the Excel file.")
    except Exception as e:
        print(f"Warning: failed to save PNG heatmaps: {e}")

    print("\n✓ Done. Open the Excel file and use the heatmaps for your submission.")


if __name__ == "__main__":
    main()
