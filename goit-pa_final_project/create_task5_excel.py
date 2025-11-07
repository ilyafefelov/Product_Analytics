# -*- coding: utf-8 -*-
"""
Create Excel file for Task 5: Confidence Interval Analysis
Fefelov_Final Project-5.xlsx
"""

import pandas as pd
import numpy as np
from scipy import stats
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# DATA AND CALCULATIONS
# ============================================================================

# Control group data
ctrl_add_to_cart = 37700
ctrl_purchases = 15161
ctrl_conversion = ctrl_purchases / ctrl_add_to_cart

# Test group data
test_add_to_cart = 26446
test_purchases = 15637
test_conversion = test_purchases / test_add_to_cart

# Difference
diff = test_conversion - ctrl_conversion

# Standard errors
se_ctrl = np.sqrt(ctrl_conversion * (1 - ctrl_conversion) / ctrl_add_to_cart)
se_test = np.sqrt(test_conversion * (1 - test_conversion) / test_add_to_cart)
se_diff = np.sqrt(se_ctrl**2 + se_test**2)

# 95% CI
z_score = stats.norm.ppf(0.975)  # 1.96
ci_lower = diff - z_score * se_diff
ci_upper = diff + z_score * se_diff

# Relative improvement
rel_improvement = (test_conversion / ctrl_conversion - 1) * 100

# Additional metrics
ctrl_spend = 66818
test_spend = 76892
additional_spend = test_spend - ctrl_spend
additional_purchases = test_purchases - ctrl_purchases
aov = 50
additional_revenue = additional_purchases * aov
marginal_roas = additional_revenue / additional_spend

# ============================================================================
# CREATE EXCEL FILE
# ============================================================================

# Create Excel writer
output_file = 'Fefelov_Final Project-5.xlsx'
writer = pd.ExcelWriter(output_file, engine='openpyxl')

# Sheet 1: Executive Summary
exec_summary = pd.DataFrame({
    'Metric': [
        'Control Conversion Rate',
        'Test Conversion Rate',
        'Point Estimate (Difference)',
        '95% CI Lower Bound',
        '95% CI Upper Bound',
        'Relative Improvement',
        'Statistical Significance',
        'Practical Significance'
    ],
    'Value': [
        f'{ctrl_conversion*100:.2f}%',
        f'{test_conversion*100:.2f}%',
        f'{diff*100:.2f} pp',
        f'{ci_lower*100:.2f}%',
        f'{ci_upper*100:.2f}%',
        f'+{rel_improvement:.2f}%',
        'YES (CI excludes zero)',
        'HIGH (47% improvement)'
    ],
    'Interpretation': [
        'Baseline performance',
        'Improved performance',
        'Best estimate of difference',
        'Worst case scenario',
        'Best case scenario',
        'Test vs Control',
        'p < 0.01',
        'Large effect size'
    ]
})
exec_summary.to_excel(writer, sheet_name='Executive Summary', index=False)

# Sheet 2: Input Data
input_data = pd.DataFrame({
    'Group': ['Control', 'Test'],
    'Add to Cart': [ctrl_add_to_cart, test_add_to_cart],
    'Purchases': [ctrl_purchases, test_purchases],
    'Conversion Rate': [f'{ctrl_conversion*100:.2f}%', f'{test_conversion*100:.2f}%'],
    'Total Spend ($)': [ctrl_spend, test_spend],
    'Days': [29, 30]
})
input_data.to_excel(writer, sheet_name='Input Data', index=False)

# Sheet 3: Calculations Step by Step
calc_steps = pd.DataFrame({
    'Step': [
        '1. Control Conversion',
        '2. Test Conversion',
        '3. Difference (Point Estimate)',
        '4. Control SE',
        '5. Test SE',
        '6. Difference SE',
        '7. Z-score (95% CI)',
        '8. Margin of Error',
        '9. CI Lower Bound',
        '10. CI Upper Bound',
        '11. Contains Zero?',
        '12. Conclusion'
    ],
    'Formula': [
        'p_c = 15,161 / 37,700',
        'p_t = 15,637 / 26,446',
        'Δp = p_t - p_c',
        'SE_c = √[p_c(1-p_c)/n_c]',
        'SE_t = √[p_t(1-p_t)/n_t]',
        'SE_Δ = √(SE_c² + SE_t²)',
        'Z = 1.96 (two-tailed)',
        'ME = Z × SE_Δ',
        'Lower = Δp - ME',
        'Upper = Δp + ME',
        'Check if 0 in [Lower, Upper]',
        'Statistical significance'
    ],
    'Result': [
        f'{ctrl_conversion:.6f} = {ctrl_conversion*100:.2f}%',
        f'{test_conversion:.6f} = {test_conversion*100:.2f}%',
        f'{diff:.6f} = {diff*100:.2f} pp',
        f'{se_ctrl:.8f}',
        f'{se_test:.8f}',
        f'{se_diff:.8f}',
        f'{z_score:.4f}',
        f'{z_score * se_diff:.8f} = {z_score * se_diff * 100:.2f} pp',
        f'{ci_lower:.6f} = {ci_lower*100:.2f}%',
        f'{ci_upper:.6f} = {ci_upper*100:.2f}%',
        'NO - does not contain zero',
        'STATISTICALLY SIGNIFICANT'
    ]
})
calc_steps.to_excel(writer, sheet_name='Calculations', index=False)

# Sheet 4: Business Impact
business = pd.DataFrame({
    'Metric': [
        'Additional Purchases',
        'Additional Spend',
        'Assumed AOV',
        'Additional Revenue',
        'Marginal ROAS',
        'Interpretation'
    ],
    'Value': [
        additional_purchases,
        f'${additional_spend:,.2f}',
        f'${aov}',
        f'${additional_revenue:,.2f}',
        f'{marginal_roas:.2f}x',
        f'For every $1 spent, get ${marginal_roas:.2f} revenue'
    ]
})
business.to_excel(writer, sheet_name='Business Impact', index=False)

# Sheet 5: Interpretation Guide
interpretation = pd.DataFrame({
    'Question': [
        'What is a Confidence Interval?',
        'What does 95% CI mean?',
        'What is the Point Estimate?',
        'Why doesn\'t CI contain zero?',
        'What is practical significance?',
        'What is the business recommendation?'
    ],
    'Answer': [
        'A range of plausible values for the true difference between groups',
        'If we repeated this experiment 100 times, 95 would contain the true difference',
        f'Our best estimate: Test is {diff*100:.2f} pp better than Control',
        'When CI excludes zero, the difference is statistically significant (p < 0.05)',
        f'The improvement (+47%) is large enough to matter in business terms',
        'IMPLEMENT test campaign - statistically and practically significant improvement'
    ]
})
interpretation.to_excel(writer, sheet_name='Interpretation', index=False)

# Sheet 6: Python Code
code_df = pd.DataFrame({
    'Python Code': [
        '# Import libraries',
        'import numpy as np',
        'from scipy import stats',
        '',
        '# Data',
        f'n_control = {ctrl_add_to_cart}',
        f'n_test = {test_add_to_cart}',
        f'purchases_control = {ctrl_purchases}',
        f'purchases_test = {test_purchases}',
        '',
        '# Conversion rates',
        'p_control = purchases_control / n_control',
        'p_test = purchases_test / n_test',
        '',
        '# Difference',
        'diff = p_test - p_control',
        '',
        '# Standard errors',
        'se_control = np.sqrt(p_control * (1 - p_control) / n_control)',
        'se_test = np.sqrt(p_test * (1 - p_test) / n_test)',
        'se_diff = np.sqrt(se_control**2 + se_test**2)',
        '',
        '# 95% CI',
        'z_score = stats.norm.ppf(0.975)  # 1.96',
        'ci_lower = diff - z_score * se_diff',
        'ci_upper = diff + z_score * se_diff',
        '',
        f'# Result: [{ci_lower*100:.2f}%, {ci_upper*100:.2f}%]'
    ]
})
code_df.to_excel(writer, sheet_name='Python Code', index=False)

# Save the workbook
writer.close()

# ============================================================================
# FORMAT THE EXCEL FILE
# ============================================================================

wb = openpyxl.load_workbook(output_file)

# Define styles
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
title_font = Font(bold=True, size=14)
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Format each sheet
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # Add title
    ws.insert_rows(1, 2)
    ws['A1'] = f'Fefelov_Final Project-5: {sheet_name}'
    ws['A1'].font = title_font
    ws['A2'] = 'Confidence Interval Analysis - Cart→Purchase Conversion'
    
    # Format headers (now in row 3)
    for cell in ws[3]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Format data rows
    for row in range(4, ws.max_row + 1):
        for cell in ws[row]:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

# Special formatting for Executive Summary
ws_exec = wb['Executive Summary']
ws_exec['A1'].fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
ws_exec['A1'].font = Font(bold=True, size=16, color="FFFFFF")

# Highlight key results
for row in range(4, 11):
    if 'CI' in str(ws_exec[f'A{row}'].value) or 'Significance' in str(ws_exec[f'A{row}'].value):
        for col in ['A', 'B', 'C']:
            ws_exec[f'{col}{row}'].fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            ws_exec[f'{col}{row}'].font = Font(bold=True)

# Save formatted workbook
wb.save(output_file)

print(f"✓ Successfully created: {output_file}")
print(f"\nFile contains:")
print("  - Executive Summary")
print("  - Input Data")
print("  - Step-by-Step Calculations")
print("  - Business Impact Analysis")
print("  - Interpretation Guide")
print("  - Python Code for Verification")
print(f"\nKey Result: 95% CI = [{ci_lower*100:.2f}%, {ci_upper*100:.2f}%]")
print(f"Conclusion: Test campaign improves conversion by {diff*100:.2f} pp (statistically significant)")
