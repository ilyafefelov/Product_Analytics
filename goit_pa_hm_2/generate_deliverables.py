"""
Генерація файлів для Домашнього завдання №2
Продукт: Інтернет-магазин електроніки (схожий на Rozetka)
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================================
# ФАЙЛ 1: Основні метрики та їх розрахунок
# ============================================================================

def create_metrics_file():
    """Створення файлу з описом основних метрик"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Основні метрики"
    
    # Заголовок
    ws['A1'] = "ТАКСОНОМІЯ ТА ОСНОВНІ МЕТРИКИ"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:D1')
    
    ws['A2'] = "Продукт: Інтернет-магазин електроніки"
    ws['A2'].font = Font(size=12, bold=True)
    ws.merge_cells('A2:D2')
    
    # Структура метрик
    data = [
        ["", "", "", ""],
        ["КАТЕГОРІЯ", "МЕТРИКА", "ФОРМУЛА РОЗРАХУНКУ", "ОПИС"],
        ["", "", "", ""],
        
        # Acquisition (Залучення)
        ["ACQUISITION", "Кількість візитів", "COUNT(DISTINCT session_id)", "Загальна кількість сесій на сайті"],
        ["ACQUISITION", "Кількість унікальних користувачів", "COUNT(DISTINCT user_id)", "Кількість унікальних відвідувачів"],
        ["ACQUISITION", "Трафік за джерелами", "COUNT(session_id) GROUP BY traffic_source", "Розподіл візитів за каналами (organic, paid, direct, social)"],
        ["ACQUISITION", "Вартість залучення (CAC)", "SUM(marketing_cost) / COUNT(DISTINCT new_customers)", "Витрати на маркетинг / кількість нових клієнтів"],
        ["", "", "", ""],
        
        # Activation (Активація)
        ["ACTIVATION", "Відсоток реєстрацій", "(COUNT(registered_users) / COUNT(visitors)) * 100", "% відвідувачів, які зареєструвались"],
        ["ACTIVATION", "Час до першої взаємодії", "AVG(first_interaction_time - landing_time)", "Середній час від заходу до першого кліку"],
        ["ACTIVATION", "Відсоток додавання до кошика", "(COUNT(add_to_cart) / COUNT(product_views)) * 100", "% переглядів товарів, що завершились додаванням"],
        ["ACTIVATION", "Глибина перегляду", "AVG(pages_per_session)", "Середня кількість сторінок за сесію"],
        ["", "", "", ""],
        
        # Retention (Утримання)
        ["RETENTION", "Retention Rate (Day 7)", "(COUNT(users_returned_day7) / COUNT(new_users)) * 100", "% користувачів, що повернулись на 7-й день"],
        ["RETENTION", "Retention Rate (Day 30)", "(COUNT(users_returned_day30) / COUNT(new_users)) * 100", "% користувачів, що повернулись на 30-й день"],
        ["RETENTION", "Частота покупок", "COUNT(orders) / COUNT(DISTINCT user_id)", "Середня кількість замовлень на користувача"],
        ["RETENTION", "Середній час між покупками", "AVG(order_date - previous_order_date)", "Середній інтервал між повторними покупками"],
        ["", "", "", ""],
        
        # Revenue (Дохід)
        ["REVENUE", "GMV (Gross Merchandise Value)", "SUM(order_total)", "Загальна вартість всіх замовлень"],
        ["REVENUE", "AOV (Average Order Value)", "SUM(order_total) / COUNT(orders)", "Середній чек замовлення"],
        ["REVENUE", "ARPU (Average Revenue Per User)", "SUM(revenue) / COUNT(DISTINCT user_id)", "Середній дохід на користувача"],
        ["REVENUE", "LTV (Customer Lifetime Value)", "AOV * avg_orders_per_user * avg_customer_lifespan", "Прогнозована цінність клієнта за весь період"],
        ["REVENUE", "Conversion Rate", "(COUNT(orders) / COUNT(sessions)) * 100", "% сесій, що завершились покупкою"],
        ["", "", "", ""],
        
        # Referral (Реферали)
        ["REFERRAL", "Відсоток користувачів з рефералами", "(COUNT(users_with_referrals) / COUNT(users)) * 100", "% користувачів, що запросили інших"],
        ["REFERRAL", "K-factor", "invites_sent * conversion_rate", "Вірусний коефіцієнт росту"],
        ["", "", "", ""],
        
        # Engagement (Залученість)
        ["ENGAGEMENT", "Час на сайті", "AVG(session_duration)", "Середня тривалість сесії"],
        ["ENGAGEMENT", "Bounce Rate", "(COUNT(single_page_sessions) / COUNT(sessions)) * 100", "% сесій з переглядом однієї сторінки"],
        ["ENGAGEMENT", "Відсоток використання пошуку", "(COUNT(search_sessions) / COUNT(sessions)) * 100", "% сесій з використанням пошуку"],
        ["ENGAGEMENT", "CTR на рекомендації", "(COUNT(recommendation_clicks) / COUNT(recommendation_views)) * 100", "% кліків по рекомендованих товарах"],
        ["", "", "", ""],
        
        # Product Performance (Ефективність товарів)
        ["PRODUCT", "Кількість переглядів товару", "COUNT(product_view_events)", "Загальна кількість переглядів"],
        ["PRODUCT", "Conversion товару", "(COUNT(product_purchases) / COUNT(product_views)) * 100", "% переглядів, що завершились покупкою"],
        ["PRODUCT", "Додавання до вішлиста", "COUNT(add_to_wishlist)", "Кількість додавань до списку бажань"],
        ["PRODUCT", "Повернення товарів", "(COUNT(returns) / COUNT(delivered_orders)) * 100", "% повернених товарів"],
    ]
    
    # Заповнення даних
    for row_idx, row_data in enumerate(data, start=3):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Форматування заголовків
            if row_idx == 5:  # Рядок з назвами колонок
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.font = Font(bold=True, color="FFFFFF")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Форматування категорій
            elif col_idx == 1 and value and value.isupper():
                cell.font = Font(bold=True, size=10)
                cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            
            # Звичайне вирівнювання
            else:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    # Налаштування ширини колонок
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 50
    
    # Додатковий аркуш з поясненнями
    ws2 = wb.create_sheet("Пояснення")
    
    explanations = [
        ["ПОЯСНЕННЯ ДО РОЗРАХУНКУ МЕТРИК", ""],
        ["", ""],
        ["Джерела даних:", ""],
        ["", "• Google Analytics / Amplitude / Mixpanel - для веб-аналітики"],
        ["", "• Внутрішня база даних - для транзакційних даних"],
        ["", "• CRM система - для даних про клієнтів"],
        ["", "• Маркетингові платформи - для витрат на рекламу"],
        ["", ""],
        ["Частота оновлення:", ""],
        ["", "• Метрики залучення та активації - в реальному часі"],
        ["", "• Метрики доходу - щоденно"],
        ["", "• Retention та LTV - щотижнево/щомісячно"],
        ["", ""],
        ["Рекомендації щодо впровадження:", ""],
        ["", "1. Налаштувати систему збору подій (tracking plan)"],
        ["", "2. Створити дашборди для моніторингу ключових метрик"],
        ["", "3. Встановити цільові значення (targets) для кожної метрики"],
        ["", "4. Регулярно аналізувати воронку конверсії"],
        ["", "5. Сегментувати користувачів для глибшого аналізу"],
    ]
    
    for row_idx, (col1, col2) in enumerate(explanations, start=1):
        ws2.cell(row=row_idx, column=1, value=col1).font = Font(bold=True if row_idx == 1 else False, size=12 if row_idx == 1 else 10)
        ws2.cell(row=row_idx, column=2, value=col2)
        if row_idx == 1:
            ws2.merge_cells(f'A{row_idx}:B{row_idx}')
    
    ws2.column_dimensions['A'].width = 25
    ws2.column_dimensions['B'].width = 70
    
    wb.save("d:\\School\\GoIT\\Courses\\Product_Analytics\\goit_pa_hm_2\\Fefelov_PA_assignment_2-1.xlsx")
    print("✓ Створено файл Fefelov_PA_assignment_2-1.xlsx")


# ============================================================================
# ФАЙЛ 2: Tracking Plan
# ============================================================================

def create_tracking_plan():
    """Створення файлу з tracking plan"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Tracking Plan"
    
    # Заголовок
    ws['A1'] = "TRACKING PLAN: ПЕРШЕ ВІДВІДУВАННЯ САЙТУ"
    ws['A1'].font = Font(size=14, bold=True)
    ws['A1'].alignment = Alignment(horizontal='center')
    ws.merge_cells('A1:F1')
    
    ws['A2'] = "Продукт: Інтернет-магазин електроніки | Сценарій: Перший візит користувача"
    ws['A2'].font = Font(size=11, italic=True)
    ws.merge_cells('A2:F2')
    
    # Заголовки таблиці
    headers = ["№", "НАЗВА ПОДІЇ", "ОПИС", "ПАРАМЕТРИ (Properties)", "ПРИКЛАД ЗНАЧЕНЬ", "КОЛИ СПРАЦЬОВУЄ"]
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Дані подій
    events = [
        [
            "1",
            "page_view",
            "Перегляд головної сторінки",
            "• page_url\n• page_title\n• referrer\n• utm_source\n• utm_medium\n• utm_campaign\n• device_type\n• browser\n• is_first_visit",
            "page_url: '/'\npage_title: 'Головна'\nreferrer: 'google.com'\nutm_source: 'google'\nutm_medium: 'cpc'\ndevice_type: 'mobile'\nbrowser: 'Chrome'\nis_first_visit: true",
            "Коли користувач завантажує головну сторінку"
        ],
        [
            "2",
            "session_start",
            "Початок сесії користувача",
            "• session_id\n• user_id\n• timestamp\n• traffic_source\n• landing_page\n• country\n• city",
            "session_id: 'ses_12345'\nuser_id: 'anon_67890'\ntraffic_source: 'organic'\nlanding_page: '/'\ncountry: 'UA'\ncity: 'Kyiv'",
            "При першому заході на сайт (автоматично)"
        ],
        [
            "3",
            "banner_view",
            "Перегляд промо-банера",
            "• banner_id\n• banner_name\n• banner_position\n• banner_type\n• promotion_name",
            "banner_id: 'promo_001'\nbanner_name: 'Black Friday'\nbanner_position: 'hero'\nbanner_type: 'seasonal'\npromotion_name: '50% Off'",
            "Коли банер з'являється у viewport користувача"
        ],
        [
            "4",
            "banner_click",
            "Клік по промо-банеру",
            "• banner_id\n• banner_name\n• click_position_x\n• click_position_y\n• destination_url",
            "banner_id: 'promo_001'\nbanner_name: 'Black Friday'\ndestination_url: '/promotions/black-friday'",
            "Коли користувач клікає на банер"
        ],
        [
            "5",
            "category_click",
            "Клік по категорії товарів",
            "• category_id\n• category_name\n• category_level\n• click_location",
            "category_id: 'cat_smartphones'\ncategory_name: 'Смартфони'\ncategory_level: '1'\nclick_location: 'header_menu'",
            "Коли користувач обирає категорію з меню"
        ],
        [
            "6",
            "search_initiated",
            "Початок пошуку",
            "• search_query\n• search_location\n• suggestions_shown\n• query_length",
            "search_query: 'iphone 15'\nsearch_location: 'header'\nsuggestions_shown: true\nquery_length: 9",
            "Коли користувач починає вводити текст у пошук"
        ],
        [
            "7",
            "search_submitted",
            "Виконання пошуку",
            "• search_query\n• results_count\n• search_time_ms\n• filters_applied",
            "search_query: 'iphone 15'\nresults_count: 47\nsearch_time_ms: 234\nfilters_applied: []",
            "Коли користувач натискає Enter або кнопку пошуку"
        ],
        [
            "8",
            "product_list_view",
            "Перегляд списку товарів",
            "• category_id\n• category_name\n• products_shown\n• sort_type\n• filter_applied\n• page_number",
            "category_name: 'Смартфони'\nproducts_shown: 24\nsort_type: 'popularity'\nfilter_applied: 'price_range'\npage_number: 1",
            "Коли завантажується сторінка зі списком товарів"
        ],
        [
            "9",
            "product_click",
            "Клік по товару",
            "• product_id\n• product_name\n• product_price\n• product_brand\n• list_position\n• list_name",
            "product_id: 'prod_12345'\nproduct_name: 'iPhone 15 Pro'\nproduct_price: 45999\nproduct_brand: 'Apple'\nlist_position: 3\nlist_name: 'category_smartphones'",
            "Коли користувач клікає на картку товару"
        ],
        [
            "10",
            "product_view",
            "Перегляд сторінки товару",
            "• product_id\n• product_name\n• product_price\n• product_brand\n• product_category\n• availability\n• discount_amount",
            "product_id: 'prod_12345'\nproduct_name: 'iPhone 15 Pro'\nproduct_price: 45999\nproduct_brand: 'Apple'\nproduct_category: 'Смартфони'\navailability: 'in_stock'\ndiscount_amount: 0",
            "Коли завантажується детальна сторінка товару"
        ],
        [
            "11",
            "add_to_cart",
            "Додавання товару в кошик",
            "• product_id\n• product_name\n• product_price\n• quantity\n• cart_total\n• source_page",
            "product_id: 'prod_12345'\nproduct_name: 'iPhone 15 Pro'\nproduct_price: 45999\nquantity: 1\ncart_total: 45999\nsource_page: 'product_page'",
            "Коли користувач натискає 'Додати в кошик'"
        ],
        [
            "12",
            "signup_modal_view",
            "Перегляд модального вікна реєстрації",
            "• modal_trigger\n• trigger_location\n• time_on_site",
            "modal_trigger: 'add_to_cart'\ntrigger_location: 'product_page'\ntime_on_site: 180",
            "Коли показується запрошення зареєструватись"
        ],
    ]
    
    # Заповнення даних
    for row_idx, event_data in enumerate(events, start=5):
        for col_idx, value in enumerate(event_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            
            # Виділення номера події
            if col_idx == 1:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='top')
            
            # Виділення назви події
            if col_idx == 2:
                cell.font = Font(bold=True, color="0066CC")
    
    # Налаштування ширини колонок
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 28
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 35
    ws.column_dimensions['F'].width = 35
    
    # Налаштування висоти рядків
    for row in range(5, 17):
        ws.row_dimensions[row].height = 100
    
    # Додатковий аркуш з методологією
    ws2 = wb.create_sheet("Методологія")
    
    methodology = [
        ["МЕТОДОЛОГІЯ ЗБОРУ ДАНИХ", ""],
        ["", ""],
        ["Інструменти трекінгу:", ""],
        ["", "• Google Analytics 4 - веб-аналітика та поведінка користувачів"],
        ["", "• Google Tag Manager - управління тегами та подіями"],
        ["", "• Amplitude/Mixpanel - продуктова аналітика"],
        ["", "• Custom event tracking - власна система логування"],
        ["", ""],
        ["Принципи іменування подій:", ""],
        ["", "• Використовувати snake_case для назв подій"],
        ["", "• Дієслова у формі дії (view, click, add, remove)"],
        ["", "• Консистентність у назвах параметрів"],
        ["", ""],
        ["Обов'язкові параметри для всіх подій:", ""],
        ["", "• timestamp - час події"],
        ["", "• user_id - ідентифікатор користувача"],
        ["", "• session_id - ідентифікатор сесії"],
        ["", "• event_id - унікальний ідентифікатор події"],
        ["", ""],
        ["Збереження даних:", ""],
        ["", "• Real-time streaming до BigQuery/Snowflake"],
        ["", "• Batch processing для історичних даних"],
        ["", "• Retention period: 25 місяців"],
        ["", ""],
        ["Контроль якості даних:", ""],
        ["", "• Автоматична валідація схеми подій"],
        ["", "• Моніторинг аномалій у потоці даних"],
        ["", "• Щотижнева перевірка точності трекінгу"],
    ]
    
    for row_idx, (col1, col2) in enumerate(methodology, start=1):
        ws2.cell(row=row_idx, column=1, value=col1).font = Font(bold=True if row_idx == 1 else False, size=12 if row_idx == 1 else 10)
        ws2.cell(row=row_idx, column=2, value=col2)
        if row_idx == 1:
            ws2.merge_cells(f'A{row_idx}:B{row_idx}')
    
    ws2.column_dimensions['A'].width = 30
    ws2.column_dimensions['B'].width = 70
    
    wb.save("d:\\School\\GoIT\\Courses\\Product_Analytics\\goit_pa_hm_2\\Fefelov_PA_assignment_2-2.xlsx")
    print("✓ Створено файл Fefelov_PA_assignment_2-2.xlsx")


# ============================================================================
# Головна функція
# ============================================================================

if __name__ == "__main__":
    print("Генерація файлів для Домашнього завдання №2...")
    print("=" * 60)
    
    create_metrics_file()
    create_tracking_plan()
    
    print("=" * 60)
    print("✓ Всі файли успішно створені!")
    print("\nФайли готові до завантаження в LMS:")
    print("  1. Fefelov_PA_assignment_2-1.xlsx - Основні метрики")
    print("  2. Fefelov_PA_assignment_2-2.xlsx - Tracking Plan")
